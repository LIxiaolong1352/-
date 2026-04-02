import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 创建Excel工作簿
wb = Workbook()
ws = wb.active
ws.title = "问卷数据"

# 设置列标题
headers = [
    "序号", "提交时间", 
    "1.性别", "2.身份", "3.运动年限", "4.每周频率",
    "5.是否受伤", "6.受伤次数", "7.损伤类型", "8.损伤情境", "9.损伤原因",
    "10.准备活动频率", "11.准备活动时长", "12.预防知识", "13.受伤后处理",
    "14.开放题建议"
]

# 写入标题行
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 设置列宽
column_widths = [6, 15, 8, 15, 12, 12, 10, 10, 20, 12, 12, 12, 12, 12, 12, 30]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[chr(64+i)].width = width

# 添加示例数据行（演示用，实际调查时删除）
example_data = [
    [1, "2024-04-01", "男", "普通爱好者", "1-3年", "3-4次", "是", "1次", "髌骨劳损", "起跳落地", "准备活动没做好", "经常做", "5-10分钟", "了解一些", "立即停止", "加强热身"],
    [2, "2024-04-01", "男", "体育学院篮球专项", "3-5年", "5次及以上", "是", "2次", "半月板损伤;韧带损伤", "急停变向", "太累/疲劳", "每次都做", "10-20分钟", "非常了解", "立即停止", "控制训练量"],
    [3, "2024-04-01", "女", "普通爱好者", "1年以下", "1-2次", "否", "", "", "", "", "偶尔做", "5分钟以内", "不太了解", "没受过伤", ""],
]

for row_idx, row_data in enumerate(example_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# 添加数据说明sheet
ws_info = wb.create_sheet("数据编码说明")

info_data = [
    ["变量名", "题目", "编码说明"],
    ["gender", "1.性别", "男=1, 女=2"],
    ["identity", "2.身份", "1=体育学院篮球专项, 2=校队成员, 3=普通爱好者"],
    ["years", "3.运动年限", "1=1年以下, 2=1-3年, 3=3-5年, 4=5年以上"],
    ["frequency", "4.每周频率", "1=1-2次, 2=3-4次, 3=5次以上"],
    ["injury", "5.是否受伤", "0=否, 1=是"],
    ["injury_num", "6.受伤次数", "1=1次, 2=2次, 3=3次以上"],
    ["injury_type", "7.损伤类型", "多选用;分隔，如：半月板损伤;韧带损伤"],
    ["situation", "8.损伤情境", "1=起跳落地, 2=急停变向, 3=防守滑步, 4=身体对抗, 5=其他"],
    ["cause", "9.损伤原因", "1=准备活动, 2=技术错误, 3=疲劳, 4=场地, 5=其他"],
    ["warmup", "10.准备活动频率", "1=每次都做, 2=经常, 3=偶尔, 4=很少/从不"],
    ["warmup_time", "11.准备活动时长", "1=5分钟内, 2=5-10分, 3=10-20分, 4=20分以上"],
    ["knowledge", "12.预防知识", "1=非常了解, 2=了解一些, 3=不太了解, 4=完全不了解"],
    ["after_injury", "13.受伤后处理", "1=停止就医, 2=简单处理继续, 3=不管继续, 4=没受过伤"],
    ["", "", ""],
    ["数据分析要点", "", ""],
    ["1. 损伤发生率", "=COUNTIF(是否受伤列,\"是\")/COUNTA(序号列)", ""],
    ["2. 群体差异", "体育专项损伤率 vs 普通爱好者损伤率", ""],
    ["3. 致因排序", "准备活动、技术错误、疲劳、场地的占比", ""],
]

for row_idx, row_data in enumerate(info_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

ws_info.column_dimensions['A'].width = 15
ws_info.column_dimensions['B'].width = 25
ws_info.column_dimensions['C'].width = 50

# 保存文件
wb.save("C:\\Users\\Administrator\\Desktop\\问卷数据录入模板.xlsx")
print("Excel模板已生成！")
